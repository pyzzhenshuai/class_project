import pandas as pd
import openpyxl


# 读取faceDR文件
data_dr = []
with open('faceDR', 'r') as file_dr:
    for line_dr in file_dr:
        if line_dr.startswith(' '):
            line_dr = line_dr.strip()
            # 提取第一列和第三列数据
            parts_dr = line_dr.split()
            column1_dr = parts_dr[0]
            column3_dr = parts_dr[2]
            # 将第一列和第三列数据添加到data_dr列表中
            data_dr.append((column1_dr, column3_dr))

# 读取faceDS文件
data_ds = []
with open('faceDS', 'r') as file_ds:
    for line_ds in file_ds:
        if line_ds.startswith(' '):
            line_ds = line_ds.strip()
            # 提取第一列和第三列数据
            parts_ds = line_ds.split()
            column1_ds = parts_ds[0]
            column3_ds = parts_ds[2]
            # 将第一列和第三列数据添加到data_ds列表中
            data_ds.append((column1_ds, column3_ds))

# 创建DataFrame对象

df_dr = pd.DataFrame(data_dr)
df_ds = pd.DataFrame(data_ds)

# 合并两个DataFrame
df_combined = pd.concat([df_dr, df_ds], ignore_index=True)

# 保存为Excel文件（保存在当前路径下）
df_combined.to_excel('combined_data.xlsx', header=False, index=False, engine='openpyxl')

# 将'female'改为0，'male'改为1
df_combined.replace({'female)': 0, 'male)': 1}, inplace=True)

# 保存为Excel文件（保存在当前路径下）
df_combined.to_excel('combined_data.xlsx', header=True, index=False, engine='openpyxl')

# 打开Excel文件
workbook = openpyxl.load_workbook('combined_data.xlsx')

# 选择第一个工作表
worksheet = workbook.active

# # 删除不存在数据
# 查找包含 'descriptor)' 的行，并删除
rows_to_delete = []
for row in worksheet.iter_rows():
    for cell in row:
        if cell.value == 'descriptor)':
            rows_to_delete.append(cell.row)

for row in reversed(rows_to_delete):
    worksheet.delete_rows(row)

# 将下面的行向上补充
max_row = worksheet.max_row
for row in range(1, max_row):
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=row+1, column=col)
        worksheet.cell(row=row, column=col, value=cell.value)

# 清除最后一行的内容
for col in range(1, worksheet.max_column + 1):
    worksheet.cell(row=max_row, column=col, value='')

# 保存修改后的Excel文件
workbook.save('combined_data.xlsx')